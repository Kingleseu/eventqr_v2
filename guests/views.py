# guests/views.py
from __future__ import annotations

import csv
import re
import uuid
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils.timezone import now
from django.urls import reverse, NoReverseMatch
from django.db.models import Q
from django.views.decorators.http import require_POST

from events.models import Event
from .models import Invite, Table
from .forms import InviteForm, TableForm, AddInviteToTableForm
import json
from django.views.decorators.csrf import csrf_exempt

# ---- Imports optionnels (PDF/Excel) ----
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
except Exception:
    load_workbook = None
    Workbook = None

try:
    from weasyprint import HTML
except Exception:
    HTML = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def _active_event(request):
    """
    Récupère l'événement actif via ?event=ID, sinon via session,
    sinon le plus récent de type 'guests'.
    """
    event_id = request.GET.get("event")
    if event_id:
        try:
            ev = Event.objects.get(pk=event_id)
            request.session["active_event_id"] = ev.pk
            return ev
        except Event.DoesNotExist:
            pass

    sid = request.session.get("active_event_id")
    if sid:
        try:
            return Event.objects.get(pk=sid)
        except Event.DoesNotExist:
            pass

    ev = Event.objects.filter(event_type="guests").order_by("-starts_at", "-id").first()
    if ev:
        request.session["active_event_id"] = ev.pk
    return ev


def _build_pdf_bytes(invite: Invite, event: Event, message: str | None) -> bytes | None:
    """
    Rend un PDF en priorité avec WeasyPrint, sinon avec ReportLab.
    """
    ctx = {"invite": invite, "message": message or "", "event": event}

    if HTML is not None:
        html = render_to_string("guests/pdf_template.html", ctx)
        buf = BytesIO()
        HTML(string=html).write_pdf(buf)
        return buf.getvalue()

    if REPORTLAB_OK:
        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        width, height = A4
        y = height - 50

        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, y, f"Invitation a {event.name}")
        y -= 20

        c.setFont("Helvetica", 10)
        try:
            c.drawString(
                50, y,
                f"{event.starts_at:%d/%m/%Y %H:%M} au {event.ends_at:%d/%m/%Y %H:%M}"
            )
        except Exception:
            pass
        y -= 30

        c.setFont("Helvetica", 12)
        prenom = getattr(invite, "prenom", "") or getattr(invite, "first_name", "")
        nom = getattr(invite, "nom", "") or getattr(invite, "last_name", "") or getattr(invite, "name", "")
        c.drawString(50, y, f"Bonjour {prenom} {nom},")
        y -= 20

        msg = (message or "Veuillez trouver ci-joint votre invitation.")
        for line in msg.splitlines() or [" "]:
            c.drawString(50, y, line)
            y -= 16
        y -= 10

        table = getattr(invite, "table", None)
        if table:
            label = getattr(table, "label", None) or getattr(table, "numero", None) or "-"
            c.drawString(50, y, f"Table : {label}")
            y -= 20

        try:
            if getattr(invite, "qr_code", None) and hasattr(invite.qr_code, "path") and invite.qr_code.path:
                c.drawImage(invite.qr_code.path, 50, max(80, y - 160),
                            width=140, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

        c.showPage()
        c.save()
        return buf.getvalue()

    return None


def _public_form_url(request, event):
    if not event:
        return ""
    try:
        url = reverse("guests:public_invite_form", kwargs={"event_id": event.pk})
        return request.build_absolute_uri(url)
    except NoReverseMatch:
        return ""


# ---------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------

@login_required
def home(request):
    ev = _active_event(request)
    invites_qs = Invite.objects.filter(event=ev) if ev else Invite.objects.none()

    invites_count = invites_qs.count()
    present_count = invites_qs.filter(present=True).count()
    tables_count = Table.objects.filter(event=ev).count() if ev else 0
    unassigned_count = invites_qs.filter(table__isnull=True).count()
    presence_rate = round((present_count / invites_count) * 100, 1) if invites_count else 0.0

    return render(request, "guests/home.html", {
        "active_event": ev,
        "events": Event.objects.filter(event_type="guests").order_by("-starts_at"),
        "invites_count": invites_count,
        "present_count": present_count,
        "tables_count": tables_count,
        "unassigned_count": unassigned_count,
        "presence_rate": presence_rate,
        "last_invites": invites_qs.order_by("-id")[:10],
        "public_form_url": _public_form_url(request, ev),
    })


@login_required
def user_settings(request):
    return render(request, "guests/param.html", {
        "active_event": _active_event(request),
        "events": Event.objects.filter(event_type="guests").order_by("-starts_at"),
    })


# ---------------------------------------------------------------------
# CRUD Invites
# ---------------------------------------------------------------------

@login_required
def add_invite(request):
    ev = _active_event(request)
    if not ev:
        messages.warning(request, "Veuillez d'abord selectionner/creer un evenement 'guests'.")
        return redirect("guests:home")
    evt_type = getattr(ev, "event_type", None) or getattr(ev, "type", None)
    if evt_type != "guests":
        alt = Event.objects.filter(event_type="guests").order_by("-starts_at", "-id").first()
        if alt:
            ev = alt
            request.session["active_event_id"] = ev.pk
            messages.info(request, f"Bascule vers l'evenement invite \"{ev.name}\".")
        else:
            messages.error(request, "Aucun evenement de type 'guests' disponible.")
            return redirect("guests:home")

    public_form_url = _public_form_url(request, ev)
    invites = Invite.objects.filter(event=ev).order_by("nom", "prenom")
    form = InviteForm(event=ev)

    if request.method == "POST":
        if "add_invite" in request.POST:
            form = InviteForm(request.POST, request.FILES, event=ev)
            if form.is_valid():
                inv = form.save(commit=False)
                inv.event = ev
                inv.user = request.user
                # Génère un couple_id si on est en mode couple
                invite_type = request.POST.get("invite_type", "individual")
                couple_id = str(uuid.uuid4()) if invite_type == "couple" else None
                if couple_id:
                    inv.couple_id = couple_id
                inv.save()

                invite_type = request.POST.get("invite_type", "individual")
                base_table = inv.table
                extra_invites = []

                if invite_type == "couple":
                    fn = (request.POST.get("couple_first_name_2") or "").strip()
                    ln = (request.POST.get("couple_last_name_2") or "").strip()
                    tele2 = (request.POST.get("couple_tele_2") or "").strip()
                    if fn or ln or tele2:
                        extra_invites.append({"prenom": fn or "Invite", "nom": ln or "", "tele": tele2})

                if invite_type == "group":
                    fn = (request.POST.get("group_first_name_2") or "").strip()
                    ln = (request.POST.get("group_last_name_2") or "").strip()
                    tele2 = (request.POST.get("group_tele_2") or "").strip()
                    if fn or ln or tele2:
                        extra_invites.append({"prenom": fn or "Invite", "nom": ln or "", "tele": tele2})

                pattern = re.compile(r"^(?P<prefix>individual|couple|group)_first_name_extra_(?P<idx>\d+)$")
                for key, val in request.POST.items():
                    m = pattern.match(key)
                    if not m:
                        continue
                    idx = m.group("idx")
                    prefix = m.group("prefix")
                    fn = (val or "").strip()
                    ln = (request.POST.get(f"{prefix}_last_name_extra_{idx}") or "").strip()
                    tele_extra = (request.POST.get(f"{prefix}_tele_extra_{idx}") or "").strip()
                    if fn or ln or tele_extra:
                        extra_invites.append({"prenom": fn or "Invite", "nom": ln or "", "tele": tele_extra})

                for data in extra_invites:
                    Invite.objects.create(
                        event=ev,
                        user=request.user,
                        nom=data["nom"],
                        prenom=data["prenom"],
                        tele=data["tele"],
                        table=base_table,
                        couple_id=couple_id if invite_type == "couple" else None,
                    )

                messages.success(request, f"{1 + len(extra_invites)} invite(s) ajoute(s).")
                return redirect(f"{request.path}?event={ev.pk}")
            messages.error(request, "Veuillez corriger le formulaire.")

        elif "import_excel" in request.POST:
            upload = request.FILES.get("excel_file")
            if not upload:
                messages.error(request, "Aucun fichier sélectionné.")
                return redirect(f"{request.path}?event={ev.pk}")

            name = upload.name.lower()
            inserted, ignored = 0, 0
            preview = []
            try:
                if name.endswith(".xlsx"):
                    if load_workbook is None:
                        messages.error(request, "L'import Excel nécessite openpyxl.")
                        return redirect(f"{request.path}?event={ev.pk}")
                    wb = load_workbook(filename=upload)
                    sheet = wb.active
                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        rows.append(row)
                elif name.endswith(".csv"):
                    text = upload.read().decode("utf-8")
                    reader = csv.reader(text.splitlines())
                    rows = list(reader)[1:]  # skip header
                else:
                    messages.error(request, "Formats acceptés : .xlsx ou .csv")
                    return redirect(f"{request.path}?event={ev.pk}")

                for row in rows:
                    title = (row[0] or "").strip() if len(row) > 0 else ""
                    prenom = (row[1] or "").strip() if len(row) > 1 else ""
                    nom = (row[2] or "").strip() if len(row) > 2 else ""
                    email = (row[3] or "").strip() if len(row) > 3 else ""
                    tele = (row[4] or "").strip() if len(row) > 4 else ""
                    couple_id = (row[5] or "").strip() if len(row) > 5 else ""
                    if nom and prenom:
                        Invite.objects.create(
                            event=ev, user=request.user,
                            nom=nom, prenom=prenom, tele=tele, email=email or None,
                            couple_id=couple_id or None
                        )
                        preview.append(f"{prenom} {nom}")
                        inserted += 1
                    else:
                        ignored += 1

                if inserted:
                    ap = ", ".join(preview[:10]) + ("..." if len(preview) > 10 else "")
                    messages.success(request, f"{inserted} invité(s) importé(s) : {ap}")
                if ignored:
                    messages.warning(request, f"{ignored} ligne(s) ignorée(s) (champs manquants).")
            except Exception as e:
                messages.error(request, f"Erreur d'import : {e}")

            return redirect(f"{request.path}?event={ev.pk}")

    return render(request, "guests/add_invite.html", {
        "event": ev,
        "active_event": ev,
        "events": Event.objects.filter(event_type="guests").order_by("-starts_at"),
        "form": form,
        "invites": invites,
        "public_form_url": public_form_url,
    })


@login_required
def update_invite(request, pk: int):
    ev = _active_event(request)
    inv = get_object_or_404(Invite, pk=pk, event=ev)
    if request.method == "POST":
        form = InviteForm(request.POST, request.FILES, instance=inv, event=ev)
        if form.is_valid():
            form.save()
            return redirect(f"{reverse('guests:add_invite')}?event={ev.pk}")
    else:
        form = InviteForm(instance=inv, event=ev)

    return render(request, "guests/add_invite.html", {
        "active_event": ev,
        "events": Event.objects.filter(event_type="guests").order_by("-starts_at"),
        "form": form,
        "invites": Invite.objects.filter(event=ev).order_by("nom", "prenom"),
        "editing": True,
        "editing_invite": inv,
    })


@login_required
def delete_invite(request, pk: int):
    ev = _active_event(request)
    inv = get_object_or_404(Invite, pk=pk, event=ev)
    if request.method == "POST":
        if getattr(inv, "couple_id", None):
            Invite.objects.filter(event=ev, couple_id=inv.couple_id).delete()
            messages.success(request, "Couple supprimé.")
        else:
            inv.delete()
            messages.success(request, "Invite supprime.")
    return redirect(f"{reverse('guests:add_invite')}?event={ev.pk}")


@login_required
def bulk_action(request):
    ev = _active_event(request)
    if request.method != "POST":
        return redirect(f"{reverse('guests:add_invite')}?event={getattr(ev, 'pk', '')}")

    action = (request.POST.get("action") or "").strip()
    ids_raw = (request.POST.get("ids") or "").strip()
    try:
        ids = [int(x) for x in ids_raw.split(",") if x.strip().isdigit()]
    except Exception:
        ids = []

    if not ids:
        messages.warning(request, "Aucune sélection.")
        return redirect(f"{reverse('guests:add_invite')}?event={getattr(ev, 'pk', '')}")

    qs = Invite.objects.filter(event=ev, pk__in=ids)
    if not qs.exists():
        messages.warning(request, "Aucun invité trouvé pour cette action.")
        return redirect(f"{reverse('guests:add_invite')}?event={getattr(ev, 'pk', '')}")

    if action == "delete":
        couple_ids = list(qs.exclude(couple_id__isnull=True).values_list("couple_id", flat=True))
        to_delete = Invite.objects.filter(event=ev).filter(Q(pk__in=ids) | Q(couple_id__in=couple_ids))
        count = to_delete.count()
        to_delete.delete()
        messages.success(request, f"{count} invité(s) supprimé(s).")
    elif action in ("present", "absent"):
        present_state = action == "present"
        updated = qs.update(present=present_state, date_scanned=now() if present_state else None)
        messages.success(request, f"{updated} invité(s) mis à jour.")
    else:
        messages.warning(request, "Action inconnue.")

    return redirect(f"{reverse('guests:add_invite')}?event={getattr(ev, 'pk', '')}")


# ---------------------------------------------------------------------
# Tables
# ---------------------------------------------------------------------
@login_required
def admin_page(request):
    ev = _active_event(request)
    if not ev:
        messages.warning(request, "Creez d'abord un evenement 'Invites'.")
        return redirect("dashboard:index")

    table_form = TableForm()
    try:
        link_form = AddInviteToTableForm(event=ev)
    except TypeError:
        link_form = AddInviteToTableForm()

    if request.method == "POST":
        if "create_table" in request.POST:
            table_form = TableForm(request.POST)
            if table_form.is_valid():
                tbl = table_form.save(commit=False)
                tbl.event = ev
                tbl.save()
                label = getattr(tbl, "label", None) or getattr(tbl, "numero", None) or tbl.pk
                messages.success(request, f"Table creee : {label}")
                return redirect("guests:admin")

        elif "generate_tables" in request.POST:
            try:
                count = int(request.POST.get("count", 0))
                capacity = int(request.POST.get("capacity", 0))
                shape = request.POST.get("shape", "round")
                
                if count > 0 and capacity > 0:
                    # Trouve le numero max actuel (si numerique)
                    existing_nums = []
                    for t in Table.objects.filter(event=ev):
                        try:
                            # Essaie "Table 1" -> 1 ou juste "1" -> 1
                            lbl = (t.label if hasattr(t, "label") else t.numero) or "0"
                            num = int(lbl.lower().replace("table", "").strip())
                            existing_nums.append(num)
                        except ValueError:
                            pass
                    
                    next_start = (max(existing_nums) if existing_nums else 0) + 1
                    
                    new_tables = []
                    for i in range(count):
                        new_tables.append(Table(
                            event=ev,
                            numero=str(next_start + i),
                            nombre_de_places=capacity,
                            shape=shape
                        ))
                    Table.objects.bulk_create(new_tables)
                    messages.success(request, f"{count} tables générées (capacité {capacity}).")
                else:
                    messages.warning(request, "Valeurs invalides.")
            except ValueError:
                messages.error(request, "Erreur de saisie.")
            return redirect("guests:admin")

        elif "add_invite_to_table" in request.POST:
            try:
                link_form = AddInviteToTableForm(request.POST, event=ev)
            except TypeError:
                link_form = AddInviteToTableForm(request.POST)

            if link_form.is_valid():
                invite = link_form.cleaned_data["invite"]
                table = link_form.cleaned_data["table"]

                if table.event_id != ev.id:
                    messages.error(request, "La table ne correspond pas a l'evenement actif.")
                else:
                    if hasattr(table, "invites"):
                        rel = table.invites
                    elif hasattr(table, "invite_set"):
                        rel = table.invite_set
                    elif hasattr(table, "guests"):
                        rel = table.guests
                    else:
                        rel = Invite.objects.filter(table=table)

                    capacity = getattr(table, "capacity", None) or getattr(table, "nombre_de_places", None)
                    count = rel.count() if hasattr(rel, "count") else Invite.objects.filter(table=table).count()

                    if capacity and count >= capacity:
                        messages.warning(request, "Cette table est complete.")
                    else:
                        invite.table = table
                        invite.save()
                        prenom = getattr(invite, "prenom", "") or getattr(invite, "first_name", "")
                        nom = getattr(invite, "nom", "") or getattr(invite, "last_name", "") or getattr(invite, "name", "")
                        label = getattr(table, "label", None) or getattr(table, "numero", None)
                        messages.success(request, f"{prenom} {nom} assigne a {label}")
                return redirect("guests:admin")

    # Récupérer les tables et les trier naturellement (1, 2, 10... et pas 1, 10, 2)
    raw_tables = list(Table.objects.filter(event=ev))
    def natural_sort_key(t):
        # Trie par numéro (int si possible), sinon alphanumérique
        label = getattr(t, "label", None) or getattr(t, "numero", "")
        # Essaie d'extraire la partie numérique
        import re
        nums = re.findall(r'\d+', str(label))
        if nums:
            return int(nums[0])
        return str(label)
    
    raw_tables.sort(key=natural_sort_key)
    tables = []
    for t in raw_tables:
        if hasattr(t, "invites"):
            rel = t.invites
        elif hasattr(t, "invite_set"):
            rel = t.invite_set
        elif hasattr(t, "guests"):
            rel = t.guests
        else:
            rel = Invite.objects.filter(table=t)

        invites_qs = rel.all() if hasattr(rel, "all") else rel
        
        # Display logic: "NOM (Num)" or just "Table Num"
        d_label = getattr(t, "numero", "")
        if getattr(t, "nom", None):
            d_label = f"{t.nom} <small class='text-muted'>({t.numero})</small>"
        else:
            d_label = f"Table {d_label}"

        tables.append({
            "obj": t,
            "display_label": d_label,
            "capacity": getattr(t, "capacity", None) or getattr(t, "nombre_de_places", None),
            "count": invites_qs.count(),
            "invites": invites_qs,
        })

    return render(request, "guests/admin_page.html", {
        "active_event": ev,
        "events": Event.objects.filter(event_type="guests").order_by("-starts_at"),
        "table_form": table_form,
        "invite_form": link_form,
        "tables": tables,
        "unassigned_guests": Invite.objects.filter(event=ev, table__isnull=True).order_by("nom", "prenom"),
    })


@csrf_exempt
@login_required
def move_guest(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Method not allowed"}, status=405)
    
    try:
        data = json.loads(request.body)
        guest_id = data.get("guest_id")
        table_id = data.get("table_id") # Can be None/null for unassign
        
        ev = _active_event(request)
        primary_guest = get_object_or_404(Invite, pk=guest_id, event=ev)
        
        # Determine target table
        if table_id:
            new_table = get_object_or_404(Table, pk=table_id, event=ev)
        else:
            new_table = None

        # Identify all guests to move (Self + Couple)
        guests_to_move = [primary_guest]
        if primary_guest.couple_id:
            # Find partner(s)
            partners = Invite.objects.filter(event=ev, couple_id=primary_guest.couple_id).exclude(pk=primary_guest.pk)
            guests_to_move.extend(partners)

        # Track involved tables for stats update
        affected_table_ids = set()
        if primary_guest.table: affected_table_ids.add(primary_guest.table.id)
        if new_table: affected_table_ids.add(new_table.id)

        # Execute Moves
        moved_ids = []
        for g in guests_to_move:
            if g.table: affected_table_ids.add(g.table.id) # Add partner's old table
            g.table = new_table
            g.save()
            moved_ids.append(g.id)
        
        # Prepare response data with updated counts
        resp_data = {
            "status": "success", 
            "message": "Guests moved.",
            "moved_guests": moved_ids
        }

        def get_table_stats(tid):
            try:
                tbl = Table.objects.get(pk=tid)
                c = Invite.objects.filter(table=tbl).count()
                cap = getattr(tbl, "capacity", None) or getattr(tbl, "nombre_de_places", None) or 0
                return {
                    "id": tbl.id,
                    "count": c,
                    "capacity": cap
                }
            except Table.DoesNotExist:
                return None

        # Return stats for ALL affected tables (Source(s) + Dest)
        # We put them in a list 'updated_tables'
        updated_tables = []
        for tid in affected_table_ids:
            stats = get_table_stats(tid)
            if stats: updated_tables.append(stats)
            
        resp_data["updated_tables"] = updated_tables
        
        return JsonResponse(resp_data)

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
def delete_table(request, table_id: int):
    ev = _active_event(request)
    tbl = get_object_or_404(Table, pk=table_id, event=ev)
    if request.method == "POST":
        tbl.delete()
        
        # Support AJAX
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
             return JsonResponse({"status": "success", "message": "Table supprimée."})
             
        messages.success(request, "Table supprimee.")
    return redirect("guests:admin")


@login_required
def edit_table(request, table_id: int):
    ev = _active_event(request)
    tbl = get_object_or_404(Table, pk=table_id, event=ev)
    
    if request.method == "POST":
        form = TableForm(request.POST, instance=tbl)
        if form.is_valid():
            t = form.save()
            
            # Support AJAX
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "success",
                    "table": {
                        "id": t.id,
                        "nom": t.nom,
                        "numero": t.numero,
                        "color": t.color,
                        "shape": t.shape,
                        "capacity": t.nombre_de_places
                    }
                })

            messages.success(request, "Table modifiee avec succes.")
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"status": "error", "errors": form.errors}, status=400)
            messages.error(request, "Erreur lors de la modification.")
    
    return redirect("guests:admin")


# ---------------------------------------------------------------------
# Invitations (PDF, Email, WhatsApp)
# ---------------------------------------------------------------------

@login_required
def envoyer_invitations(request):
    ev = _active_event(request)
    invites = Invite.objects.filter(event=ev).select_related("table").order_by("nom", "prenom")

    if request.method == "POST":
        invite_id = request.POST.get("invites")
        message_personnalise = (request.POST.get("message") or "").strip()
        inv = get_object_or_404(Invite, pk=invite_id, event=ev)

        if not getattr(inv, "qr_code", None):
            return HttpResponse("Aucun QR code enregistre pour cet invite.")

        pdf_bytes = _build_pdf_bytes(inv, ev, message_personnalise)

        if "send_email" in request.POST:
            from django.core.mail import EmailMessage
            if not getattr(inv, "email", ""):
                messages.error(request, "Aucune adresse email pour cet invite.")
                return redirect("guests:envoyer_invitations")

            email = EmailMessage(
                subject=f"Invitation a {ev.name}",
                body=message_personnalise or "Veuillez trouver ci-joint votre invitation.",
                to=[getattr(inv, "email", "")],
            )
            if pdf_bytes:
                prenom = getattr(inv, "prenom", "") or getattr(inv, "first_name", "")
                nom = getattr(inv, "nom", "") or getattr(inv, "last_name", "") or getattr(inv, "name", "")
                email.attach(f"Invitation_{nom}_{prenom}.pdf", pdf_bytes, "application/pdf")
            else:
                messages.warning(
                    request,
                    "PDF non genere (WeasyPrint/ReportLab non installes). Envoi sans piece jointe."
                )
            email.send()
            messages.success(request, "Invitation envoyee par email.")
            return redirect("guests:envoyer_invitations")

        if "send_whatsapp" in request.POST:
            tel = getattr(inv, "tele", "") or getattr(inv, "telephone", "")
            url = f"http://wa.me/{tel}?text={message_personnalise}"
            return HttpResponse(f'<script>window.location="{url}";</script>')

    return render(request, "guests/message.html", {
        "active_event": ev,
        "events": Event.objects.filter(event_type="guests").order_by("-starts_at"),
        "invites": invites,
    })


# ---------------------------------------------------------------------
# Presences / Export CSV
# ---------------------------------------------------------------------

@login_required
def liste_presences(request):
    ev = _active_event(request)
    invites = Invite.objects.filter(event=ev).order_by("nom", "prenom")
    return render(request, "guests/liste_presences.html", {
        "active_event": ev,
        "events": Event.objects.filter(event_type="guests").order_by("-starts_at"),
        "invites": invites,
    })


@login_required
def export_presences_csv(request):
    ev = _active_event(request)
    invites = Invite.objects.filter(event=ev).order_by("nom", "prenom")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="presences.csv"'
    writer = csv.writer(response)
    writer.writerow(["Nom", "Prenom", "Telephone", "Table", "Presence", "Date de scan"])
    for i in invites:
        table = getattr(i, "table", None)
        label = getattr(table, "label", None) or getattr(table, "numero", None) or "-"
        writer.writerow([
            getattr(i, "nom", "") or getattr(i, "last_name", "") or getattr(i, "name", ""),
            getattr(i, "prenom", "") or getattr(i, "first_name", ""),
            getattr(i, "tele", "") or getattr(i, "telephone", ""),
            label,
            ("Present" if getattr(i, "present", False) else "Absent"),
            (getattr(i, "date_scanned", None).strftime("%Y-%m-%d %H:%M") if getattr(i, "date_scanned", None) else ""),
        ])
    return response


@login_required
def export_invites(request):
    """
    Exporte la liste d'invités courante (CSV par défaut, XLSX si ?fmt=xlsx et openpyxl dispo).
    """
    ev = _active_event(request)
    invites = Invite.objects.filter(event=ev).order_by("nom", "prenom")
    fmt = (request.GET.get("fmt") or "csv").lower()

    headers = ["Title", "Prenom", "Nom", "Email", "Telephone", "Couple_ID", "Present", "Tag"]

    if fmt == "xlsx" and Workbook is not None:
        wb = Workbook()
        sh = wb.active
        sh.title = "Invites"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # Add Headers
        sh.append(headers)
        for cell in sh[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add Data
        for i in invites:
            sh.append([
                getattr(i, "titre", "") or "",
                i.prenom or "",
                i.nom or "",
                getattr(i, "email", "") or "",
                getattr(i, "tele", "") or "",
                i.couple_id or "",
                "Oui" if getattr(i, "present", False) else "Non",
                getattr(i, "tag", "") if hasattr(i, "tag") else "",
            ])
            
        # Apply borders to all data cells
        for row in sh.iter_rows(min_row=2, max_row=1 + len(invites), min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # Auto-adjust column widths
        for col in sh.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sh.column_dimensions[column].width = adjusted_width
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="invites.xlsx"'
        wb.save(response)
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename=\"invites.csv\"'
    w = csv.writer(response)
    w.writerow(headers)
    for i in invites:
        w.writerow([
            getattr(i, "titre", "") or "",
            i.prenom or "",
            i.nom or "",
            getattr(i, "email", "") or "",
            getattr(i, "tele", "") or "",
            i.couple_id or "",
            "Oui" if getattr(i, "present", False) else "Non",
            getattr(i, "tag", "") if hasattr(i, "tag") else "",
        ])
    return response


# ---------------------------------------------------------------------
# Scan QR minimal
# ---------------------------------------------------------------------

@login_required
def valider_qr_code(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Methode non autorisee."})
    
    ev = _active_event(request)
    qr_data = (request.POST.get("qr_data") or "").strip()
    if not qr_data:
        return JsonResponse({"status": "error", "message": "Aucune donnee QR recue."})
    try:
        data = dict(part.strip().split(": ", 1) for part in qr_data.split(" | "))
        prenom = data.get("Prénom", "").strip()
        nom = data.get("Nom", "").strip()
        inv = get_object_or_404(Invite, prenom__iexact=prenom, nom__iexact=nom, event=ev)
        if getattr(inv, "present", False):
            return JsonResponse({"status": "warning", "message": f"{inv} deja present."})
        inv.present = True
        inv.date_scanned = now()
        inv.save()
        return JsonResponse({"status": "success", "message": f"{inv} marque present."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Erreur: {e}"})


# ---------------------------------------------------------------------
# Visual Table Plan
# ---------------------------------------------------------------------

@login_required
def table_plan(request):
    ev = _active_event(request)
    if not ev:
        messages.warning(request, "Sélectionnez un événement d'abord.")
        return redirect("guests:admin")
    
    tables = Table.objects.filter(event=ev)
    return render(request, "guests/table_plan.html", {
        "active_event": ev,
        "tables": tables,
    })

@csrf_exempt
@login_required
def save_table_positions(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST required"}, status=405)
    
    try:
        data = json.loads(request.body)
        positions = data.get("positions", [])
        
        updated_count = 0
        for item in positions:
            t_id = item.get("id")
            x = float(item.get("x", 0))
            y = float(item.get("y", 0))
            
            # Update efficiently
            Table.objects.filter(pk=t_id).update(x_pos=x, y_pos=y)
            updated_count += 1
            
        return JsonResponse({"status": "success", "count": updated_count})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


# ---------------------------------------------------------------------
# Modele d'import (Excel/CSV)
# ---------------------------------------------------------------------

@login_required
def download_invites_template(request):
    headers = ["Title", "Prenom", "Nom", "Email", "Telephone", "Couple_ID", "AddGuestsAllowed", "Tag"]
    instructions = [
        "Laisser vide si inconnu.\n(Ex: M., Mme, Dr.)",
        "Prénom de l'invité.\nObligatoire si Nom absent.",
        "Nom de famille.\nObligatoire si Prénom absent.",
        "Adresse email.\nOptionnel.",
        "Numéro de téléphone.\nRecommandé pour WhatsApp/SMS.",
        "Identifiant unique pour lier des invités ensemble (Couple/Famille).\nMême valeur pour chaque personne du groupe.\n(Ex: COUPle-001)",
        "Nombre d'invités supplémentaires autorisés.\n(Ex: 0, 1, 2, ou 'unlimited').",
        "Étiquettes pour catégoriser.\n(Ex: VIP, Famille)."
    ]
    sample_rows = [
        ["M.", "Jean", "Dupont", "jean.dupont@mail.com", "0601020304", "", "unlimited", "VIP"],
        ["Mme", "Marie", "Curie", "marie.curie@mail.com", "0605060708", "COUPLE-001", "none", "Famille"],
        ["M.", "Pierre", "Curie", "", "0605060709", "COUPLE-001", "none", "Famille"],
    ]

    if Workbook is not None:
        wb = Workbook()
        sh = wb.active
        sh.title = "Invites"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        instruction_font = Font(italic=True, color="555555", size=9)
        instruction_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # --- Row 1: Headers ---
        sh.append(headers)
        for cell in sh[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- Row 2: Instructions ---
        sh.append(instructions)
        for cell in sh[2]:
            cell.font = instruction_font
            cell.alignment = instruction_alignment
            cell.border = thin_border

        # --- Row 3+: Data ---
        for row in sample_rows:
            sh.append(row)
            
        # Apply borders to all data cells (starting from row 3)
        total_rows = 2 + len(sample_rows)
        for row in sh.iter_rows(min_row=3, max_row=total_rows, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # --- Column Widths ---
        # Fixed widths to make it look like the template
        column_widths = [15, 25, 25, 30, 20, 30, 25, 20]
        for i, width in enumerate(column_widths, start=1):
            col_letter = sh.cell(row=1, column=i).column_letter
            sh.column_dimensions[col_letter].width = width

        # Freeze the top 2 rows
        sh.freeze_panes = "A3"

        tips = wb.create_sheet("Instructions")
        tips.append(["Colonne", "Description"])
        tips.append(["Title", "Optionnel. Ex: M., Mme, Dr."])
        tips.append(["Prenom", "Obligatoire si Nom absent (au moins l'un des deux)."])
        tips.append(["Nom", "Obligatoire si Prenom absent (au moins l'un des deux)."])
        tips.append(["Email", "Optionnel (utile pour envoi email)."])
        tips.append(["Telephone", "Recommandé (WhatsApp/SMS)."])
        tips.append(["Couple_ID", "Optionnel. Même valeur pour relier deux personnes (couple/famille)."])
        tips.append(["AddGuestsAllowed", "Optionnel. Valeurs : unlimited / none / 0 / 2 ..."])
        tips.append(["Tag", "Optionnel. Ex: VIP, Famille."])
        tips.append([])
        tips.append(["Conseils"])
        tips.append(["- Gardez la première ligne pour les en-têtes."])
        tips.append(["- Pour un couple, saisissez la même valeur dans Couple_ID pour les deux lignes."])
        tips.append(["- Le numéro de téléphone est la donnée principale pour l'envoi de QR via WhatsApp/SMS."])

        # Style Instructions header
        for cell in tips[1]:
            cell.font = Font(bold=True)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="modele_invites.xlsx"'
        wb.save(response)
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="modele_invites.csv"'
    w = csv.writer(response)
    w.writerow(headers)
    for row in sample_rows:
        w.writerow(row)
    return response


@login_required
def search_invites(request):
    """
    Retourne une liste JSON d'invites filtrés par 'q' (nom/prenom/telephone)
    et optionnellement par 'event' (id d'evenement).
    """
    q = (request.GET.get("q") or "").strip()
    event_id = (request.GET.get("event") or request.session.get("active_event_id"))

    qs = Invite.objects.all()
    if event_id:
        qs = qs.filter(event_id=event_id)

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(tele__icontains=q)
        )

    qs = qs.order_by("nom", "prenom")[:20]

    data = []
    for i in qs:
        table_label = ""
        try:
            table_label = i.table.label or ""
        except Exception:
            table_label = ""

        data.append({
            "id": i.id,
            "nom": i.nom or "",
            "prenom": i.prenom or "",
            "tele": getattr(i, "tele", "") or "",
            "table": table_label,
            "present": bool(getattr(i, "present", False)),
            "detail_url": f"/guests/invites/{i.id}/",
        })
    return JsonResponse(data, safe=False)


# ---------------------------------------------------------------------
# Studio de Création (Invitation Builder No-Code)
# ---------------------------------------------------------------------

from .models import InvitationTemplate

@login_required
def invitation_studio(request, event_id: int):
    """
    Interface de design Wizard 4-Étapes (No-Code Builder).
    """
    ev = get_object_or_404(Event, pk=event_id, event_type="guests")
    
    # Assurer qu'un template existe
    template, created = InvitationTemplate.objects.get_or_create(event=ev)
    
    # Passer la liste des invités pour l'étape 4 (vrais liens)
    invites = Invite.objects.filter(event=ev).exclude(code__isnull=True).exclude(code="").order_by("nom", "prenom")
    
    return render(request, "guests/builder/studio.html", {
        "event": ev,
        "template": template,
        "invites": invites,
        "invites_count": invites.count(),
    })

@csrf_exempt
@login_required
def save_template(request, event_id: int):
    """
    Sauvegarde AJAX envoyée depuis GrapesJS.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    
    try:
        ev = get_object_or_404(Event, pk=event_id, event_type="guests")
        template, _ = InvitationTemplate.objects.get_or_create(event=ev)
        
        data = json.loads(request.body)
        template.html_content = data.get("html", "")
        template.css_content = data.get("css", "")
        template.components_data = data.get("components", "[]") # Optional JSON components
        template.styles_data = data.get("styles", "[]") # Optional JSON styles
        template.save()
        
        return JsonResponse({"status": "success", "message": "Sauvegarde réussie !"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ---------------------------------------------------------------------
# Vue Côté Invité (Mini-site web généré)
# ---------------------------------------------------------------------

def guest_invitation_view(request, code: str):
    """
    Lien magique envoyé à l'invité.
    ex: /i/ad4f-1234-5678/
    """
    invite = get_object_or_404(Invite, code=code)
    ev = invite.event
    
    try:
        template = ev.invitation_template
    except InvitationTemplate.DoesNotExist:
        return HttpResponse("Oups, aucune invitation digitale n'a été créée pour cet événement.", status=404)

    # Remplacement des variables dynamiques du builder.
    html = template.html_content
    # Exemple: L'organisateur tape "[prenom]" dans son bloc texte.
    html = html.replace("[prenom]", f"{getattr(invite, 'prenom', '')}")
    html = html.replace("[nom]", f"{getattr(invite, 'nom', '')}")
    html = html.replace("[nom_evenement]", f"{ev.name}")
    
    context = {
        "invite": invite,
        "event": ev,
        "template": template,
        "html_rendered": html
    }
    
    return render(request, "guests/public/invitation_viewer.html", context)
